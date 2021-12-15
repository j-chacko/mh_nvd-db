# Load an Excel file and then retrieve the column letter and row number
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Required to dump the data to the output file
import json

import argparse
import os


def validate_file(f):
    if not os.path.exists(f):
        # Argparse uses the ArgumentTypeError to give a rejection message like:
        # error: argument input: x does not exist
        raise argparse.ArgumentTypeError("{0} does not exist".format(f))
    return f


inputFile = ''
outputFile = ''
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Search the NVD DB and return CVE's")
    parser.add_argument("-i", "--input", dest="inputFile", required=True, type=validate_file,
                        help="input file", metavar="FILE")
    parser.add_argument("-o", "--output", dest="outputFile", required=True, help="output file", metavar="FILE")
    args = parser.parse_args()
    inputFile = args.inputFile
    outputFile = args.outputFile

wb = load_workbook(filename=inputFile)
ws = wb.active

myList = []

# Get the number of rows and number of columns in Excel for the loop
lastColumn = len(list(ws.columns))
lastRow = len(list(ws.rows))

# Create a dictionary for each row and add it the list, i.e. a list of dictionaries
for row in range(1, lastRow + 1):
    myDict = {}
    for column in range(1, lastColumn + 1):
        columnLetter = get_column_letter(column)
        if row > 1:
            myDict[ws[columnLetter + str(1)].value] = ws[columnLetter + str(row)].value
    myList.append(myDict)

# Once we have the list of dictionaries, parse it and create the JSON file
data = json.dumps(myList, sort_keys=True, indent=4, default=str)
with open(outputFile, 'w', encoding='utf-8') as f:
    f.write(data)


# Reference:
# - https://codeigo.com/python/convert-excel-spreadsheet-to-json
# - https://github.com/AccentureTVM/Python-Nmap-XML-to-CSV/blob/master/nmapxmltocsv.py
# - https://stackoverflow.com/questions/11875770/how-to-overcome-datetime-datetime-not-json-serializable
